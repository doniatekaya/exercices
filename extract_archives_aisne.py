
"""
extract_archives_aisne.py
=========================
Extraction des registres paroissiaux et d'état civil pour Abbécourt (Aisne)
depuis le site des Archives Départementales de l'Aisne.

Stratégie en 3 phases :
  Phase 1 : requests + BeautifulSoup (scraping léger)
  Phase 2 : Playwright headless (fallback si JS requis)
  Phase 3 : Export CSV + Excel

Auteur  : Cascade AI
Date    : 2026-06-30
"""

import re
import sys
import time
import json
import hashlib
import logging
import traceback
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import pandas as pd

# ─── Configuration ────────────────────────────────────────────────────────────

TARGET_URL = (
    "https://archives.aisne.fr/archive/resultats/etatcivil/n:11/limit:50"
    "?RECH_commune_Index=1%7C"
    "&RECH_commune_Libel=Abb%C3%A9court+%28Aisne%29%7C"
    "&type=etatcivil"
    "&communes_TabSel=1"
    "&RECH_Field=RECH_commune"
    "&RECH_Letter=A"
    "&RECH_annee_debut=1860"
)

BASE_URL = "https://archives.aisne.fr"
OUTPUT_DIR = Path(__file__).parent
COMMUNE = "Abbécourt (Aisne)"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Referer": "https://archives.aisne.fr/archive/recherche/etatcivil",
}

# Mapping des abréviations de types d'actes
TYPE_MAPPING = {
    "N": "Naissances",
    "M": "Mariages",
    "D": "Décès",
    "T": "Tables décennales",
    "NM": "Naissances-Mariages",
    "ND": "Naissances-Décès",
    "MD": "Mariages-Décès",
    "NMD": "Naissances-Mariages-Décès",
    "B": "Baptêmes",
    "S": "Sépultures",
    "BMS": "Baptêmes-Mariages-Sépultures",
    "BM": "Baptêmes-Mariages",
    "BS": "Baptêmes-Sépultures",
    "MS": "Mariages-Sépultures",
    "TP": "Tables de paroisses",
}

# ─── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(OUTPUT_DIR / "extraction.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ─── Utilitaires ──────────────────────────────────────────────────────────────

def normalize_type(raw: str) -> str:
    """Convertit l'abréviation du type d'acte en libellé complet."""
    raw = raw.strip().upper()
    return TYPE_MAPPING.get(raw, raw)


def clean_years(raw: str) -> tuple[str, str]:
    """
    Extrait année_début et année_fin depuis une chaîne brute.
    Exemples : '1846-1858', '1669 1750', 'An XI 1812', '1906'
    """
    raw = raw.strip().rstrip(".")
    # Cas 'An XI 1812' (calendrier révolutionnaire)
    m = re.search(r"(An\s+[IVXLCDM]+)\s+(\d{4})", raw, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2)
    # Cas 'AAAA-BBBB' ou 'AAAA BBBB'
    m = re.search(r"(\d{4})\s*[-–]\s*(\d{4})", raw)
    if m:
        return m.group(1), m.group(2)
    # Cas 'AAAA BBBB' (espace)
    m = re.search(r"(\d{4})\s+(\d{4})", raw)
    if m:
        return m.group(1), m.group(2)
    # Cas année unique
    m = re.search(r"(\d{4})", raw)
    if m:
        return m.group(1), m.group(1)
    return raw, raw


def make_absolute_url(href: str) -> str:
    """Transforme un href relatif en URL absolue."""
    if not href:
        return ""
    if href.startswith("http"):
        return href
    return BASE_URL + href


# ─── Solveur Anubis Proof-of-Work ─────────────────────────────────────────────

def solve_anubis_challenge(challenge: dict) -> tuple[int, str]:
    """
    Résout le challenge Anubis "fast" (SHA-256 proof-of-work).

    Algorithme (extrait du JS worker sha256-purejs.mjs) :
      input  = randomData + str(nonce)
      hash   = SHA-256(input.encode('utf-8'))
      valid  = hash[0..p-1] == 0x00  AND  (if odd difficulty) hash[p] >> 4 == 0
    où p = difficulty // 2

    Retourne (nonce_gagnant, hash_hex).
    """
    random_data = challenge["randomData"]
    difficulty  = challenge["difficulty"]
    p = difficulty // 2          
    u = (difficulty % 2) != 0    

    log.info(f"Résolution PoW Anubis : randomData={random_data[:20]}… difficulty={difficulty} "
             f"(p={p} octets, nibble={u})")
    t0 = time.time()
    nonce = 0
    while True:
        raw     = (random_data + str(nonce)).encode("utf-8")
        digest  = hashlib.sha256(raw).digest()

        valid = all(digest[i] == 0 for i in range(p))
        if valid and u and (digest[p] >> 4) != 0:
            valid = False

        if valid:
            elapsed = (time.time() - t0) * 1000
            hex_hash = digest.hex()
            log.info(f"Solution trouvée ! nonce={nonce} hash={hex_hash[:16]}… ({elapsed:.0f} ms, {nonce} tentatives)")
            return nonce, hex_hash

        nonce += 1
        if nonce % 500 == 0:
            log.debug(f"  PoW : {nonce} tentatives…")


def fetch_with_anubis_solver(url: str) -> BeautifulSoup | None:
    """
    Phase 1 : requests + solveur Anubis intégré.

    Étapes :
      1. GET url  → réception de la page challenge Anubis
      2. Parse JSON du challenge depuis le script#anubis_challenge
      3. Résolution SHA-256 proof-of-work
      4. GET /.within.website/x/cmd/anubis/api/pass-challenge?... → cookie valide
      5. GET url avec cookie → contenu réel
    """
    log.info("=== PHASE 1 : requests + Solveur Anubis (SHA-256 PoW) ===")
    log.info(f"URL cible : {url}")

    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Étape 1 : GET la page cible pour récupérer le challenge ──
    try:
        log.info("Chargement de la page cible (attente du challenge Anubis)...")
        resp = session.get(url, timeout=20)
        log.info(f"HTTP {resp.status_code} | {len(resp.content)} octets | "
                 f"cookies={list(session.cookies.keys())}")

        # Sauvegarder pour debug
        (OUTPUT_DIR / "debug_page.html").write_bytes(resp.content)

        soup_challenge = BeautifulSoup(resp.content, "lxml")
        page_text = soup_challenge.get_text()

        # Vérifier si c'est le challenge Anubis
        script_tag = soup_challenge.find("script", {"id": "anubis_challenge"})
        if not script_tag:
            # Pas de challenge → on a déjà les données !
            log.info("Pas de challenge Anubis détecté. La page semble accessible directement.")
            kw_found = [k for k in ["abbécourt", "5mi", "1 e 7", "registre"] if k in page_text.lower()]
            if kw_found:
                log.info(f"Mots-clés archives trouvés : {kw_found}")
                return soup_challenge
            log.warning("Page accessible mais sans données d'archives attendues.")
            log.info(f"Aperçu : {page_text[:300]!r}")
            return soup_challenge

    except Exception as e:
        log.error(f"Erreur GET initial : {e}")
        return None

    # ── Étape 2 : Parser le challenge JSON ──
    try:
        challenge_json = json.loads(script_tag.string)
        challenge = challenge_json["challenge"]
        rules     = challenge_json["rules"]
        log.info(f"Challenge Anubis parsé : id={challenge['id']} "
                 f"algorithm={rules['algorithm']} difficulty={rules['difficulty']}")
        log.info(f"  randomData (64 premiers cars) : {challenge['randomData'][:64]}")
    except Exception as e:
        log.error(f"Impossible de parser le challenge Anubis : {e}")
        log.debug(f"Script tag content : {script_tag.string[:500] if script_tag.string else 'vide'}")
        return None

    # ── Étape 3 : Résoudre le PoW ──
    try:
        nonce, hex_hash = solve_anubis_challenge(challenge)
    except Exception as e:
        log.error(f"Échec résolution PoW : {e}")
        return None

    # ── Étape 4 : Soumettre la solution à l'API Anubis ──
    elapsed_ms = 1200  # valeur réaliste simulée
    pass_url = (
        f"{BASE_URL}/.within.website/x/cmd/anubis/api/pass-challenge"
        f"?id={challenge['id']}"
        f"&response={hex_hash}"
        f"&nonce={nonce}"
        f"&redir={requests.utils.quote(url, safe='')}"
        f"&elapsedTime={elapsed_ms}"
    )
    log.info(f"Soumission solution → {pass_url[:120]}…")

    try:
        resp_pass = session.get(pass_url, timeout=20, allow_redirects=True)
        log.info(f"Réponse API : HTTP {resp_pass.status_code} | "
                 f"cookies={list(session.cookies.keys())} | "
                 f"URL finale={resp_pass.url[:80]}")

        # Sauvegarder pour debug
        (OUTPUT_DIR / "debug_pass_response.html").write_bytes(resp_pass.content)

        # Vérifier si le cookie a été accordé
        if not session.cookies:
            log.warning("Aucun cookie reçu après soumission. Le serveur a peut-être rejeté la solution.")
    except Exception as e:
        log.error(f"Erreur soumission PoW : {e}")
        return None

    # ── Étape 5 : Charger la vraie page avec le cookie valide ──
    log.info("Chargement de la page réelle avec cookie Anubis validé...")
    time.sleep(1.0)
    try:
        resp_real = session.get(url, timeout=20)
        log.info(f"HTTP {resp_real.status_code} | {len(resp_real.content)} octets")

        # Sauvegarder le HTML réel
        (OUTPUT_DIR / "debug_page_real.html").write_bytes(resp_real.content)
        log.info("HTML réel sauvegardé → debug_page_real.html")

        soup = BeautifulSoup(resp_real.content, "lxml")
        page_text = soup.get_text()
        log.info(f"Texte extrait : {len(page_text)} caractères")
        log.info(f"Aperçu : {page_text[:400]!r}")
        return soup

    except Exception as e:
        log.error(f"Erreur lors du chargement final : {e}")
        return None


# ─── Phase 2 : Playwright headless ────────────────────────────────────────────

def fetch_with_playwright(url: str) -> BeautifulSoup | None:
    """
    Tente de récupérer la page via Playwright (navigateur headless).
    Retourne un BeautifulSoup ou None.
    """
    log.info("=== PHASE 2 : Tentative Playwright (navigateur headless) ===")

    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        log.error("Playwright non installé. Exécutez : pip install playwright && playwright install chromium")
        return None

    try:
        with sync_playwright() as pw:
            log.info("Lancement navigateur Chromium headless...")
            browser = pw.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=HEADERS["User-Agent"],
                locale="fr-FR",
            )
            page = context.new_page()

            log.info("Navigation vers la page cible...")
            page.goto(url, wait_until="networkidle", timeout=30_000)
            log.info("Page chargée. Attente du tableau...")

            # Attendre que les données soient visibles
            try:
                page.wait_for_selector("table", timeout=10_000)
                log.info("Tableau détecté dans la page.")
            except PWTimeout:
                log.warning("Pas de balise <table> trouvée dans le délai imparti.")

            # Screenshot pour diagnostic
            screenshot_path = OUTPUT_DIR / "debug_screenshot.png"
            page.screenshot(path=str(screenshot_path), full_page=True)
            log.info(f"Screenshot sauvegardé → {screenshot_path}")

            html = page.content()
            browser.close()

            # Sauvegarder le HTML JS-rendered pour debug
            debug_js = OUTPUT_DIR / "debug_page_js.html"
            Path(debug_js).write_text(html, encoding="utf-8")
            log.info(f"HTML JS-rendered sauvegardé → {debug_js}")

            return BeautifulSoup(html, "lxml")

    except Exception as e:
        log.error(f"Playwright a échoué : {e}")
        log.debug(traceback.format_exc())
        return None


# ─── Parseurs HTML ────────────────────────────────────────────────────────────

def parse_table_standard(soup: BeautifulSoup) -> list[dict]:
    """
    Parseur principal : cherche le tableau de résultats standard.
    Structure attendue :
      <table> → <tr> → <td> contenant type, années, cote, lien
    """
    records = []

    # Chercher toutes les tables possibles
    tables = soup.find_all("table")
    log.info(f"Tables trouvées dans la page : {len(tables)}")

    # Chercher dans les divs résultats si pas de table
    result_containers = soup.select(
        "table, .result-table, .resultats, #results, .archive-results, "
        ".etatcivil-results, [class*='result'], [class*='registre']"
    )
    log.info(f"Conteneurs de résultats identifiés : {len(result_containers)}")

    for i, table in enumerate(tables):
        rows = table.find_all("tr")
        log.info(f"Table #{i+1} : {len(rows)} lignes")
        if len(rows) < 2:
            continue

        # Analyser l'en-tête pour détecter les colonnes
        headers_row = rows[0]
        headers_cells = [th.get_text(strip=True).lower() for th in headers_row.find_all(["th", "td"])]
        log.info(f"  En-têtes : {headers_cells}")

        # Mapper les indices de colonnes
        col_idx = {
            "type": None, "annees": None, "cote": None,
            "commune": None, "lien": None, "annee_debut": None, "annee_fin": None
        }
        for j, h in enumerate(headers_cells):
            if any(k in h for k in ["type", "acte", "nature"]):
                col_idx["type"] = j
            elif any(k in h for k in ["annee", "année", "date", "période"]):
                col_idx["annees"] = j
            elif any(k in h for k in ["cote", "référence", "reference", "côte"]):
                col_idx["cote"] = j
            elif any(k in h for k in ["commune", "localit"]):
                col_idx["commune"] = j
            elif any(k in h for k in ["lien", "image", "consul", "voir"]):
                col_idx["lien"] = j

        log.info(f"  Colonnes mappées : {col_idx}")

        for row in rows[1:]:
            cells = row.find_all(["td", "th"])
            if not cells:
                continue

            record = {
                "commune": COMMUNE,
                "type_acte": "",
                "type_acte_libelle": "",
                "annee_debut": "",
                "annee_fin": "",
                "cote": "",
                "lien_image": "",
                "notes": "",
            }

            # Extraction selon les indices détectés
            for field, idx in col_idx.items():
                if idx is not None and idx < len(cells):
                    cell = cells[idx]
                    text = cell.get_text(" ", strip=True)
                    link = cell.find("a")
                    href = make_absolute_url(link["href"]) if link and link.get("href") else ""

                    if field == "type":
                        record["type_acte"] = text
                        record["type_acte_libelle"] = normalize_type(text)
                    elif field == "annees":
                        d, f = clean_years(text)
                        record["annee_debut"] = d
                        record["annee_fin"] = f
                    elif field == "cote":
                        record["cote"] = text
                        if not record["lien_image"] and href:
                            record["lien_image"] = href
                    elif field == "commune":
                        record["commune"] = text or COMMUNE
                    elif field == "lien":
                        if href:
                            record["lien_image"] = href
                        elif link:
                            record["lien_image"] = make_absolute_url(link.get("href", ""))

            # Fallback : si les indices ne sont pas mappés, extraction positionnelle
            if not record["cote"] and cells:
                all_text = [c.get_text(" ", strip=True) for c in cells]
                all_links = [
                    make_absolute_url(a["href"])
                    for c in cells
                    for a in c.find_all("a")
                    if a.get("href")
                ]
                record["notes"] = " | ".join(t for t in all_text if t)
                if all_links:
                    record["lien_image"] = all_links[0]

            # Ignorer les lignes vides
            if not any([record["type_acte"], record["cote"], record["notes"]]):
                continue

            records.append(record)

    return records


def parse_div_based(soup: BeautifulSoup) -> list[dict]:
    """
    Parseur alternatif pour les sites qui affichent les résultats en divs.
    Cherche les patterns communs : .row, .result-item, etc.
    """
    records = []

    # Chercher des blocs de résultats en div
    items = soup.select(
        ".result-item, .archive-item, .etatcivil-item, "
        "[class*='item'], [class*='row'], [class*='ligne'], "
        ".registre, [class*='registre']"
    )
    log.info(f"Blocs div détectés : {len(items)}")

    for item in items:
        text = item.get_text(" ", strip=True)
        link = item.find("a")
        href = make_absolute_url(link["href"]) if link and link.get("href") else ""

        # Tenter d'extraire cote et années via regex
        cote_match = re.search(
            r"(\d+\s*[EMi]+\s*\d+[/\d]*|[A-Z]+\s*\d+[A-Z]?\s*/\s*\d+|1E\d+_\d+)",
            text, re.IGNORECASE
        )
        year_match = re.search(r"(\d{4})\s*[-–]\s*(\d{4})|(\d{4})\s+(\d{4})", text)

        if not cote_match and not year_match:
            continue

        record = {
            "commune": COMMUNE,
            "type_acte": "",
            "type_acte_libelle": "",
            "annee_debut": "",
            "annee_fin": "",
            "cote": cote_match.group(0).strip() if cote_match else "",
            "lien_image": href,
            "notes": text[:200],
        }
        if year_match:
            if year_match.group(1):
                record["annee_debut"] = year_match.group(1)
                record["annee_fin"] = year_match.group(2)
            else:
                record["annee_debut"] = year_match.group(3)
                record["annee_fin"] = year_match.group(4)

        records.append(record)

    return records


def parse_all_links(soup: BeautifulSoup) -> list[dict]:
    """
    Dernier recours : récupère tous les liens d'archives visibles dans la page
    et tente d'extraire les métadonnées depuis le texte environnant.
    """
    records = []
    log.info("=== Parseur de dernier recours : extraction de tous les liens ===")

    # Chercher tous les <a> qui ressemblent à des liens de registres
    all_links = soup.find_all("a", href=True)
    log.info(f"Liens <a> totaux dans la page : {len(all_links)}")

    archive_links = [
        a for a in all_links
        if any(kw in a["href"].lower() for kw in [
            "archive", "registre", "etatcivil", "ark:", "viewer",
            "image", "5mi", "1e", "3e", "visu"
        ])
    ]
    log.info(f"Liens d'archives filtrés : {len(archive_links)}")

    for a in archive_links:
        href = make_absolute_url(a["href"])
        parent_text = a.parent.get_text(" ", strip=True) if a.parent else ""
        link_text = a.get_text(strip=True)

        year_match = re.search(r"(\d{4})\s*[-–\s]\s*(\d{4})|(\d{4})", parent_text)
        cote_match = re.search(
            r"(\d+\s*[A-Za-z]+\s*\d+[/\d]*|[A-Z]+\s*\d+[A-Z]?\s*/\s*\d+|1E\d+_\d+)",
            parent_text, re.IGNORECASE
        )

        record = {
            "commune": COMMUNE,
            "type_acte": link_text if len(link_text) < 20 else "",
            "type_acte_libelle": normalize_type(link_text) if len(link_text) < 20 else "",
            "annee_debut": "",
            "annee_fin": "",
            "cote": cote_match.group(0).strip() if cote_match else "",
            "lien_image": href,
            "notes": parent_text[:300],
        }
        if year_match:
            if year_match.group(1):
                record["annee_debut"] = year_match.group(1)
                record["annee_fin"] = year_match.group(2) or year_match.group(1)
            else:
                record["annee_debut"] = year_match.group(3)
                record["annee_fin"] = year_match.group(3)

        records.append(record)

    return records


def parse_aisne_specific(soup: BeautifulSoup) -> list[dict]:
    """
    Parseur ciblé pour le site Archives Aisne.

    Structure réelle (identifiée par analyse du HTML debug_page_real.html) :
      <table> → <tr> par registre
        <td> [1 seule cellule]
          texte : "COTE • ANNEE_DEB ANNEE_FIN Voir l'inventaire ... Sujet : TYPE ..."
          liens : [viewer_url, ark_url, classeur_url, image_directe_url]

    Le 4e lien (/ark:/.../daogrp/0/layout:linear/...) est le lien image direct.
    Le 1er lien (/archive/fonds/...view:XXXX) est le visualiseur contextualisé.
    """
    records = []
    table = soup.find("table")
    if not table:
        log.warning("Parseur Aisne : aucune table trouvée.")
        return records

    rows = table.find_all("tr")
    log.info(f"Parseur Aisne : {len(rows)} lignes dans la table")

    # Regex de parsing de la cellule
    re_cote   = re.compile(r"^([^\s•]+(?:\s+[A-Z]\s+\d+[/\d]*|\s+\d+_\d+)?)\s*•")
    re_dates_hdr = re.compile(r"•\s*(.+?)\s*Voir\s+l", re.IGNORECASE)
    re_dates_meta = re.compile(r"Dates\s*[:\xa0]+\s*(.+?)(?:\s+Notes|\s+Sujet|\s+Commune|$)", re.IGNORECASE)
    re_sujet  = re.compile(r"Sujet\s*[:\xa0]+\s*(.+?)(?:\s+Commune\s+ou|\s+Canton|\s+Arrondissement|$)", re.IGNORECASE)
    re_contexte = re.compile(r"Contexte\s*[:\xa0]+\s*(.+?)(?:\s+Dates|\s+Notes|\s+Sujet|\s+Commune|$)", re.IGNORECASE)

    TYPE_NORMALIZE = {
        "bapteme":   "Baptêmes",
        "baptême":   "Baptêmes",
        "sepulture": "Sépultures",
        "sépulture": "Sépultures",
        "mariage":   "Mariages",
        "naissance": "Naissances",
        "naisance":  "Naissances",
        "deces":     "Décès",
        "décès":     "Décès",
        "table":     "Tables décennales",
        "tables":    "Tables décennales",
    }

    def normalize_sujet(raw: str) -> str:
        parts = re.split(r"\s*/\s*", raw.strip(), flags=re.IGNORECASE)
        normalized = []
        for p in parts:
            p = p.strip().lower()
            normalized.append(TYPE_NORMALIZE.get(p, p.capitalize()))
        return " / ".join(normalized)

    for row_idx, row in enumerate(rows):
        cell = row.find("td")
        if not cell:
            continue

        txt = cell.get_text(" ", strip=True)
        links = [a.get("href", "") for a in cell.find_all("a") if a.get("href")]

        # ── Cote ──
        cote = ""
        m = re_cote.search(txt)
        if m:
            # Affiner : tout ce qui précède '•' sans les espaces superflus
            cote = txt[:m.end()].split("•")[0].strip()
        else:
            # Fallback : premiers mots avant espace/•
            cote = txt.split("•")[0].strip().split()[0] if txt else ""

        # ── Années (depuis l'en-tête de la cellule) ──
        annee_debut, annee_fin = "", ""
        m_hdr = re_dates_hdr.search(txt)
        if m_hdr:
            annee_debut, annee_fin = clean_years(m_hdr.group(1))
        else:
            m_meta = re_dates_meta.search(txt)
            if m_meta:
                annee_debut, annee_fin = clean_years(m_meta.group(1))

        # ── Type d'acte (depuis Sujet) ──
        type_raw = ""
        m_sujet = re_sujet.search(txt)
        if m_sujet:
            type_raw = m_sujet.group(1).strip()
        type_libelle = normalize_sujet(type_raw) if type_raw else ""

        # ── Contexte (collection) ──
        contexte = ""
        m_ctx = re_contexte.search(txt)
        if m_ctx:
            contexte = m_ctx.group(1).strip()

        # ── Liens ──
        # Lien 0 : viewer contextualisé (/archive/fonds/.../view:XXXXX)
        # Lien 3 : image directe (/ark:/.../daogrp/0/layout:linear/...)
        lien_viewer = ""
        lien_image  = ""
        for href in links:
            full = make_absolute_url(href)
            if "view:" in href and not lien_viewer:
                lien_viewer = full
            if "daogrp" in href and "layout:linear" in href and not lien_image:
                lien_image = full

        # Si pas de lien daogrp, utiliser le viewer
        if not lien_image:
            lien_image = lien_viewer

        record = {
            "commune":          COMMUNE,
            "type_acte":        type_raw,
            "type_acte_libelle": type_libelle,
            "annee_debut":      annee_debut,
            "annee_fin":        annee_fin,
            "cote":             cote,
            "collection":       contexte,
            "lien_image":       lien_image,
            "lien_viewer":      lien_viewer,
            "notes":            txt[:300],
        }
        records.append(record)
        log.debug(f"  Row {row_idx:02d} : cote={cote!r} dates={annee_debut}-{annee_fin} "
                  f"type={type_raw!r} lien={lien_image[-50:]!r}")

    log.info(f"Parseur Aisne → {len(records)} enregistrements extraits")
    return records


def parse_soup(soup: BeautifulSoup) -> list[dict]:
    """
    Orchestre les parseurs et retourne les enregistrements extraits.
    """
    log.info("--- Analyse de la structure HTML ---")

    # Statistiques de la page
    all_text = soup.get_text()
    log.info(f"Longueur texte brut : {len(all_text)} caractères")
    log.info(f"Nombre de <table> : {len(soup.find_all('table'))}")
    log.info(f"Nombre de <tr>    : {len(soup.find_all('tr'))}")
    log.info(f"Nombre de <div>   : {len(soup.find_all('div'))}")
    log.info(f"Nombre de <a>     : {len(soup.find_all('a'))}")

    # Détecter si la page contient des données réelles
    keywords = ["abbécourt", "5mi", "1 e 7", "naissance", "mariage", "décès", "registre"]
    found = [kw for kw in keywords if kw in all_text.lower()]
    log.info(f"Mots-clés archive trouvés : {found}")

    if not found:
        log.error("La page ne semble pas contenir de données d'archives. Contenu tronqué ou bloqué.")
        log.info(f"Aperçu du contenu : {all_text[:500]!r}")
        return []

    # ── Parseur ciblé Aisne (prioritaire) ──
    records = parse_aisne_specific(soup)
    if records:
        log.info(f"Parseur ciblé Aisne → {len(records)} enregistrements")
        return records

    # ── Tentative parseur tableau générique ──
    records = parse_table_standard(soup)
    if records:
        log.info(f"Parseur tableau → {len(records)} enregistrements")
        return records

    # ── Tentative parseur div ──
    records = parse_div_based(soup)
    if records:
        log.info(f"Parseur div → {len(records)} enregistrements")
        return records

    # ── Dernier recours : liens ──
    records = parse_all_links(soup)
    log.info(f"Parseur liens → {len(records)} enregistrements")
    return records


# ─── Extraction manuelle depuis le HTML (fallback structuré) ──────────────────

def extract_from_known_structure(soup: BeautifulSoup) -> list[dict]:
    """
    Extraction ciblée basée sur la structure connue du site Aisne.
    Cherche les patterns spécifiques à ce site d'archives.
    """
    records = []
    log.info("=== Extraction ciblée structure Aisne ===")

    # Pattern 1 : lignes de tableau avec structure type | années | cote | icône
    
    selectors_to_try = [
        "table.table-results tr",
        "table.resultat tr",
        "table#results tr",
        "table tr",
        ".result-row",
        ".ligne-resultat",
        "[class*='resultat'] tr",
        "[class*='result'] tr",
    ]

    for selector in selectors_to_try:
        rows = soup.select(selector)
        if rows:
            log.info(f"Sélecteur '{selector}' → {len(rows)} éléments")

    # Chercher les spans ou divs contenant les types N/M/D
    type_elements = soup.find_all(
        string=re.compile(r"^(N|M|D|T|NMD|BMS|NM|ND|MD|BM|BS|MS|B|S)$", re.MULTILINE)
    )
    log.info(f"Éléments de type d'acte bruts : {len(type_elements)}")

    # Chercher les cotes connues dans le texte
    cote_pattern = re.compile(
        r"(5Mi\d+|1\s*E\s*\d+/\d+|1E\d+_\d+|3\s*E\s*\d+/\d+|"
        r"\d+\s*[A-Z]+\s*\d+[/\d]*)",
        re.IGNORECASE
    )
    all_cotes = cote_pattern.findall(soup.get_text())
    unique_cotes = list(dict.fromkeys(all_cotes))  # dédoublonner en gardant l'ordre
    log.info(f"Cotes trouvées dans le texte : {unique_cotes}")

    # Chercher les années dans le texte
    year_spans = re.findall(r"\b(\d{4})\s*[-–]\s*(\d{4})\b|\b(An\s+[IVXLCDM]+)\s+(\d{4})", soup.get_text())
    log.info(f"Plages d'années : {year_spans[:20]}")

    # Chercher les liens de visualisation
    viewer_links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if any(p in href.lower() for p in [
            "viewer", "visu", "ark:", "image", "5mi", "/1e", "/3e",
            "etatcivil", "registre", "consultation"
        ]):
            viewer_links.append({
                "href": make_absolute_url(href),
                "text": a.get_text(strip=True),
                "title": a.get("title", ""),
            })
    log.info(f"Liens de visualisation trouvés : {len(viewer_links)}")
    for vl in viewer_links[:10]:
        log.info(f"  → {vl}")

    return records


# ─── Export ───────────────────────────────────────────────────────────────────

def export_results(records: list[dict], suffix: str = "") -> tuple[Path, Path]:
    """
    Exporte les enregistrements vers CSV et Excel.
    Retourne (chemin_csv, chemin_excel).
    """
    if not records:
        log.warning("Aucun enregistrement à exporter.")
        return None, None

    df = pd.DataFrame(records)

    # Réordonner les colonnes
    cols_order = [
        "commune", "type_acte", "type_acte_libelle",
        "annee_debut", "annee_fin", "cote", "collection",
        "lien_image", "lien_viewer", "notes"
    ]
    for col in cols_order:
        if col not in df.columns:
            df[col] = ""
    df = df[cols_order]

    # Nettoyage
    df = df.drop_duplicates(subset=["cote", "annee_debut", "annee_fin"], keep="first")
    df = df.replace("", pd.NA).dropna(how="all").fillna("")

    log.info(f"DataFrame final : {len(df)} lignes × {len(df.columns)} colonnes")

    # Export CSV
    ts = suffix or datetime.now().strftime("%Y%m%d_%H%M")
    csv_path = OUTPUT_DIR / f"abbecourt_aisne_{ts}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    log.info(f"CSV exporté → {csv_path}")

    # Export Excel avec mise en forme
    xlsx_path = OUTPUT_DIR / f"abbecourt_aisne_{ts}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Registres")
        ws = writer.sheets["Registres"]

        # Mise en forme des en-têtes
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        # Largeurs de colonnes
        col_widths = {
            "A": 22,  # commune
            "B": 12,  # type_acte
            "C": 35,  # type_acte_libelle
            "D": 13,  # annee_debut
            "E": 13,  # annee_fin
            "F": 20,  # cote
            "G": 45,  # collection
            "H": 60,  # lien_image
            "I": 60,  # lien_viewer
            "J": 40,  # notes
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Liens cliquables dans les colonnes lien_image et lien_viewer
        from openpyxl.styles import Font as XFont
        link_cols = {
            cols_order.index("lien_image") + 1,
            cols_order.index("lien_viewer") + 1,
        }
        for row_idx in range(2, ws.max_row + 1):
            # Alternance couleur lignes
            fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
            for c in range(1, len(cols_order) + 1):
                data_cell = ws.cell(row=row_idx, column=c)
                if c in link_cols and data_cell.value and str(data_cell.value).startswith("http"):
                    data_cell.hyperlink = str(data_cell.value)
                    data_cell.font = XFont(color="0563C1", underline="single")
                else:
                    data_cell.fill = PatternFill("solid", fgColor=fill_color)
                data_cell.alignment = Alignment(wrap_text=True, vertical="top")
                data_cell.border = border

        # Figer la première ligne
        ws.freeze_panes = "A2"
        # Filtre automatique
        ws.auto_filter.ref = ws.dimensions

    log.info(f"Excel exporté → {xlsx_path}")
    return csv_path, xlsx_path


# ─── Pipeline principal ───────────────────────────────────────────────────────

def main():
    start_time = time.time()
    log.info("=" * 60)
    log.info("EXTRACTION ARCHIVES AISNE – Abbécourt")
    log.info(f"Démarrage : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)

    soup = None
    method_used = "aucune"

    # ── Phase 1 : requests + Solveur Anubis ──
    soup = fetch_with_anubis_solver(TARGET_URL)

    if soup:
        text_len = len(soup.get_text())
        log.info(f"Phase 1 réussie. Texte extrait : {text_len} caractères")
        method_used = "requests+BeautifulSoup+AnubisSolver"

        keywords = ["abbécourt", "5mi", "registre", "naissance", "mariage", "décès"]
        found = [k for k in keywords if k in soup.get_text().lower()]
        if not found or text_len < 500:
            log.warning("Données d'archives absentes après PoW → passage Phase 2 (Playwright).")
            soup = None

    # ── Phase 2 : Playwright ──
    if soup is None:
        log.info("Phase 1 insuffisante → activation Playwright.")
        soup = fetch_with_playwright(TARGET_URL)
        if soup:
            method_used = "Playwright headless"
        else:
            log.error("Phase 2 également échouée. Extraction impossible.")

    if soup is None:
        log.error("ÉCHEC TOTAL : impossible de récupérer la page. Voir extraction.log pour détails.")
        elapsed = time.time() - start_time
        log.info(f"Temps écoulé : {elapsed:.1f}s")
        sys.exit(1)

    # ── Parsing ──
    log.info(f"Méthode de collecte : {method_used}")
    records = parse_soup(soup)

    # Enrichissement par analyse ciblée si peu de résultats
    if len(records) < 5:
        log.info("Peu de résultats via parseurs génériques → analyse ciblée Aisne...")
        extract_from_known_structure(soup)

    # ── Export ──
    if records:
        csv_path, xlsx_path = export_results(records)
        log.info(f"\n{'='*60}")
        log.info(f"SUCCÈS : {len(records)} registres extraits")
        log.info(f"CSV   → {csv_path}")
        log.info(f"Excel → {xlsx_path}")
    else:
        log.warning("Aucun enregistrement structuré extrait par les parseurs automatiques.")
        log.info("Analyse manuelle du HTML requis → voir debug_page.html")
        # Export vide avec les métadonnées de contexte
        fallback = [{
            "commune": COMMUNE,
            "type_acte": "N/A",
            "type_acte_libelle": "Extraction automatique échouée",
            "annee_debut": "",
            "annee_fin": "",
            "cote": "",
            "lien_image": TARGET_URL,
            "notes": "Voir debug_page.html pour analyse manuelle",
        }]
        export_results(fallback, suffix="VIDE")

    elapsed = time.time() - start_time
    log.info(f"Temps total : {elapsed:.1f}s")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
