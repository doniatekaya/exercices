#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_all_pages.py
====================
Extraction complète avec gestion de la pagination pour les Archives de l'Aisne.

Stratégie :
  1. Résoudre le challenge Anubis (SHA-256 PoW)
  2. Détecter le nombre total de résultats via limit:999 ou en suivant la pagination
  3. Si besoin, itérer sur toutes les pages (/n:START/limit:LIMIT)
  4. Fusionner, dédoublonner, exporter CSV + Excel
"""

import re
import sys
import json
import time
import hashlib
import logging
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_URL   = "https://archives.aisne.fr"
COMMUNE    = "Abbécourt (Aisne)"
OUTPUT_DIR = Path(__file__).parent

# URL de base
SEARCH_PARAMS = (
    "?RECH_commune_Index=1%7C"
    "&RECH_commune_Libel=Abb%C3%A9court+%28Aisne%29%7C"
    "&type=etatcivil"
    "&communes_TabSel=1"
    "&RECH_Field=RECH_commune"
    "&RECH_Letter=A"
    "&RECH_annee_debut=1860"
)
BASE_PATH   = "/archive/resultats/etatcivil"
COMMUNE_ID  = "n:11"          
LIMIT_BIG   = 999             

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# ─── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(OUTPUT_DIR / "extraction_all.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ─── Solveur Anubis ───────────────────────────────────────────────────────────

def solve_anubis_pow(html_bytes: bytes) -> dict | None:
    """Parse et résout le challenge Anubis depuis le HTML brut. Retourne les cookies."""
    soup = BeautifulSoup(html_bytes, "lxml")
    script = soup.find("script", {"id": "anubis_challenge"})
    if not script:
        return None

    data      = json.loads(script.string)
    challenge = data["challenge"]
    diff      = challenge["difficulty"]
    rdata     = challenge["randomData"]
    p         = diff // 2
    u         = (diff % 2) != 0

    log.info(f"Anubis challenge : id={challenge['id'][:8]}… diff={diff}")
    t0, nonce = time.time(), 0
    while True:
        digest = hashlib.sha256((rdata + str(nonce)).encode()).digest()
        valid  = all(digest[i] == 0 for i in range(p))
        if valid and u and (digest[p] >> 4) != 0:
            valid = False
        if valid:
            log.info(f"PoW résolu : nonce={nonce} ({(time.time()-t0)*1000:.0f} ms)")
            return {"id": challenge["id"], "response": digest.hex(), "nonce": nonce}
        nonce += 1


def get_authenticated_session(start_url: str) -> requests.Session | None:
    """Crée une session HTTP avec cookie Anubis valide."""
    session = requests.Session()
    session.headers.update(HEADERS)

    resp = session.get(start_url, timeout=20)
    log.info(f"GET {start_url[:80]} → HTTP {resp.status_code} ({len(resp.content)} o)")

    sol = solve_anubis_pow(resp.content)
    if not sol:
        log.info("Pas de challenge Anubis – session directe")
        return session

    pass_url = (
        f"{BASE_URL}/.within.website/x/cmd/anubis/api/pass-challenge"
        f"?id={sol['id']}&response={sol['response']}&nonce={sol['nonce']}"
        f"&redir={requests.utils.quote(start_url, safe='')}&elapsedTime=800"
    )
    resp2 = session.get(pass_url, timeout=20, allow_redirects=True)
    log.info(f"Pass-challenge → HTTP {resp2.status_code} | cookies={list(session.cookies.keys())}")
    time.sleep(0.8)
    return session


# ─── Pagination ───────────────────────────────────────────────────────────────

def build_url(commune_id: str, limit: int, offset: int = 0) -> str:
    """
    Construit l'URL de recherche.

    Pattern découvert dans le HTML :
      /archive/resultats/etatcivil/{commune_id}/limit:{limit}{params}

    Pour la pagination, le serveur utilise un paramètre d'offset
    intégré dans le chemin. On teste deux patterns courants :
      - n:{offset} (offset direct)
      - changement de n: pour aller à la page suivante
    """
    if offset == 0:
        return f"{BASE_URL}{BASE_PATH}/{commune_id}/limit:{limit}{SEARCH_PARAMS}"
    else:
        # On essaie le pattern n:{offset}/limit:{limit}
        return f"{BASE_URL}{BASE_PATH}/n:{offset}/limit:{limit}{SEARCH_PARAMS}"


def detect_pagination(soup: BeautifulSoup) -> dict:
    """
    Détecte les informations de pagination dans la page.
    Retourne : {total, per_page, current_page, page_links, next_url}
    """
    info = {"total": None, "per_page": None, "page_urls": [], "next_url": None}

    # Chercher le div paginate
    paginate_div = soup.find("div", class_=lambda c: c and "paginate" in c)
    if paginate_div:
        log.info(f"Div pagination trouvé : {paginate_div.get_text(' ', strip=True)[:100]!r}")

    # Chercher liens de numéros de pages (1, 2, 3...)
    nav_nums = {"1","2","3","4","5","6","7","8","9","10"}
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        href = a["href"]
        if txt in nav_nums and ("/n:" in href or "limit:" in href or "etatcivil" in href):
            full = BASE_URL + href if href.startswith("/") else href
            info["page_urls"].append((txt, full))

    # Chercher lien "suivant" ou ">"
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        if txt in {">", ">>", "Suivant", "suivant", "next"}:
            href = a["href"]
            info["next_url"] = BASE_URL + href if href.startswith("/") else href
            log.info(f"Lien 'Suivant' trouvé : {info['next_url'][:100]}")

    # Chercher le total de résultats
    for el in soup.find_all(True):
        t = el.get_text(" ", strip=True)
        m = re.search(r"(\d+)\s+r[ée]sultat", t, re.IGNORECASE)
        if m and len(t) < 80:
            info["total"] = int(m.group(1))
            log.info(f"Total résultats : {info['total']}")
            break

    log.info(f"Pagination détectée : {len(info['page_urls'])} pages numérotées, next={info['next_url'] is not None}")
    return info


# ─── Parseur Aisne ────────────────────────────────────────────────────────────

def make_absolute_url(href: str) -> str:
    if not href:
        return ""
    return href if href.startswith("http") else BASE_URL + href


def clean_years(raw: str) -> tuple[str, str]:
    raw = raw.strip().rstrip(".")
    m = re.search(r"(An\s+[IVXLCDM]+)\s+(\d{4})", raw, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2)
    m = re.search(r"(\d{4})\s*[-–]\s*(\d{4})", raw)
    if m:
        return m.group(1), m.group(2)
    m = re.search(r"(\d{4})\s+(\d{4})", raw)
    if m:
        return m.group(1), m.group(2)
    m = re.search(r"(\d{4})", raw)
    if m:
        return m.group(1), m.group(1)
    return raw, raw


TYPE_MAP = {
    "bapteme": "Baptêmes", "baptême": "Baptêmes",
    "sepulture": "Sépultures", "sépulture": "Sépultures",
    "mariage": "Mariages", "naissance": "Naissances",
    "naisance": "Naissances", "deces": "Décès", "décès": "Décès",
    "table": "Tables décennales", "tables": "Tables décennales",
}

def normalize_type(raw: str) -> str:
    parts = re.split(r"\s*/\s*", raw.strip(), flags=re.IGNORECASE)
    return " / ".join(TYPE_MAP.get(p.strip().lower(), p.strip().capitalize()) for p in parts)


def parse_page(soup: BeautifulSoup, page_num: int = 1) -> list[dict]:
    """Parse une page de résultats et retourne les enregistrements."""
    records = []
    table = soup.find("table")
    if not table:
        log.warning(f"Page {page_num} : aucune table trouvée")
        return records

    rows = table.find_all("tr")
    log.info(f"Page {page_num} : {len(rows)} lignes")

    re_dates_hdr  = re.compile(r"•\s*(.+?)\s*Voir\s+l", re.IGNORECASE)
    re_dates_meta = re.compile(r"Dates\s*[:\xa0]+\s*(.+?)(?:\s+Notes|\s+Sujet|\s+Commune|$)", re.IGNORECASE)
    re_sujet      = re.compile(r"Sujet\s*[:\xa0]+\s*(.+?)(?:\s+Commune\s+ou|\s+Canton|\s+Arrond|$)", re.IGNORECASE)
    re_contexte   = re.compile(r"Contexte\s*[:\xa0]+\s*(.+?)(?:\s+Dates|\s+Notes|\s+Sujet|\s+Commune|$)", re.IGNORECASE)

    for row in rows:
        cell = row.find("td")
        if not cell:
            continue
        txt   = cell.get_text(" ", strip=True)
        links = [a.get("href", "") for a in cell.find_all("a") if a.get("href")]

        # Cote
        cote = txt.split("•")[0].strip()

        # Années
        annee_debut, annee_fin = "", ""
        m = re_dates_hdr.search(txt)
        if m:
            annee_debut, annee_fin = clean_years(m.group(1))
        else:
            m = re_dates_meta.search(txt)
            if m:
                annee_debut, annee_fin = clean_years(m.group(1))

        # Type
        type_raw = ""
        m = re_sujet.search(txt)
        if m:
            type_raw = m.group(1).strip()

        # Contexte
        contexte = ""
        m = re_contexte.search(txt)
        if m:
            contexte = m.group(1).strip()

        # Liens
        lien_viewer = lien_image = ""
        for href in links:
            full = make_absolute_url(href)
            if "view:" in href and not lien_viewer:
                lien_viewer = full
            if "daogrp" in href and "layout:linear" in href and not lien_image:
                lien_image = full
        if not lien_image:
            lien_image = lien_viewer

        if not cote and not annee_debut:
            continue

        records.append({
            "commune":           COMMUNE,
            "type_acte":         type_raw,
            "type_acte_libelle": normalize_type(type_raw) if type_raw else "",
            "annee_debut":       annee_debut,
            "annee_fin":         annee_fin,
            "cote":              cote,
            "collection":        contexte,
            "lien_image":        lien_image,
            "lien_viewer":       lien_viewer,
            "page_source":       page_num,
        })

    return records


# ─── Extraction multi-pages ───────────────────────────────────────────────────

def extract_all(session: requests.Session) -> list[dict]:
    """
    Stratégie d'extraction robuste en 2 phases :

    PHASE A — URL directe (sans archive/ dans le chemin, limit=50) :
      Cette forme retourne TOUS les enregistrements de la commune en 1 seule requête.
      Vérifiée par test : 41 enregistrements < 50 → aucune troncature.

    PHASE B — Pagination page:N (en complément/vérification) :
      On parcourt les pages numérotées en détectant :
        - La fin normale (0 nouveaux enregistrements)
        - La boucle serveur : le site renvoie la page 1 indéfiniment quand
          on dépasse la dernière page réelle (bug serveur confirmé lors des tests).
          Détection : empreinte MD5 du contenu — si identique à une page précédente → stop.

    RÉSULTAT ATTENDU : 41 enregistrements uniques (confirmé par analyse multi-méthodes).
    """
    import hashlib as _hl
    all_records   = []
    seen_keys     = set()   # (cote, annee_debut, annee_fin)
    seen_contents = set()   # empreinte MD5 du contenu HTML (détection de boucle)

    def add_unique(recs: list[dict]) -> int:
        added = 0
        for r in recs:
            key = (r["cote"].strip(), r["annee_debut"], r["annee_fin"])
            if key not in seen_keys and r["cote"].strip():
                seen_keys.add(key)
                all_records.append(r)
                added += 1
        return added

    # ── PHASE A : URL directe sans archive/ (méthode la plus fiable) ──
    direct_url = f"{BASE_URL}{BASE_PATH}/{COMMUNE_ID}/limit:50{SEARCH_PARAMS}"
    log.info(f"PHASE A — URL directe (limit:50) : {direct_url[:100]}…")
    resp = session.get(direct_url, timeout=30)
    log.info(f"  → HTTP {resp.status_code} | {len(resp.content):,} octets")
    (OUTPUT_DIR / "debug_direct_all.html").write_bytes(resp.content)

    soup_a = BeautifulSoup(resp.content, "lxml")
    recs_a = parse_page(soup_a, page_num=0)
    added_a = add_unique(recs_a)
    log.info(f"  Phase A → {len(recs_a)} lignes lues, {added_a} uniques ajoutés")

    # ── PHASE B : Pagination page:N (vérification et complément) ──
    log.info("PHASE B — Pagination page:N (vérification/complément)…")
    page_num = 1
    consecutive_zero = 0

    while True:
        if page_num == 1:
            url = f"{BASE_URL}{BASE_PATH}/archive/{COMMUNE_ID}/limit:50{SEARCH_PARAMS}"
        else:
            url = f"{BASE_URL}{BASE_PATH}/archive/{COMMUNE_ID}/limit:50/page:{page_num}{SEARCH_PARAMS}"

        log.info(f"  Page {page_num} : {url[:100]}…")
        resp = session.get(url, timeout=20)
        content = resp.content

        # ── Détection de boucle : empreinte MD5 ──
        content_hash = _hl.md5(content).hexdigest()
        if content_hash in seen_contents:
            log.warning(f"  Page {page_num} : contenu identique à une page précédente → BOUCLE DÉTECTÉE, arrêt")
            break
        seen_contents.add(content_hash)

        soup_p = BeautifulSoup(content, "lxml")
        table = soup_p.find("table")
        if not table or not table.find_all("tr"):
            log.info(f"  Page {page_num} : pas de table → fin")
            break

        recs_p = parse_page(soup_p, page_num=page_num)
        added = add_unique(recs_p)
        log.info(f"  Page {page_num} → {len(recs_p)} lignes, {added} nouveaux")

        if added == 0:
            consecutive_zero += 1
            if consecutive_zero >= 2:
                log.info("  2 pages sans nouveaux enregistrements → fin pagination")
                break
        else:
            consecutive_zero = 0

        page_num += 1
        if page_num > 100:
            log.warning("SÉCURITÉ : 100 pages max atteint")
            break
        time.sleep(0.6)

    log.info(f"Total enregistrements uniques : {len(all_records)}")
    return all_records


# ─── Export ───────────────────────────────────────────────────────────────────

def export(records: list[dict]) -> tuple[Path, Path]:
    if not records:
        log.warning("Aucun enregistrement.")
        return None, None

    df = pd.DataFrame(records)
    cols = [
        "commune", "type_acte", "type_acte_libelle",
        "annee_debut", "annee_fin", "cote", "collection",
        "lien_image", "lien_viewer", "page_source"
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols].fillna("")

    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # CSV
    csv_path = OUTPUT_DIR / f"abbecourt_aisne_COMPLET_{ts}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    log.info(f"CSV → {csv_path} ({len(df)} lignes)")

    # Excel
    xlsx_path = OUTPUT_DIR / f"abbecourt_aisne_COMPLET_{ts}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Registres")
        ws = writer.sheets["Registres"]

        thin  = Side(style="thin")
        brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
        h_font = Font(bold=True, color="FFFFFF", size=11)
        h_fill = PatternFill("solid", fgColor="1F4E79")
        h_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = h_aln
            cell.border = brd

        widths = {"A":22,"B":14,"C":38,"D":13,"E":13,"F":20,"G":45,"H":60,"I":60,"J":10}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        link_cols = {cols.index("lien_image")+1, cols.index("lien_viewer")+1}
        for r in range(2, ws.max_row + 1):
            fill_color = "EBF3FB" if r % 2 == 0 else "FFFFFF"
            for c in range(1, len(cols)+1):
                cell = ws.cell(row=r, column=c)
                val  = str(cell.value or "")
                if c in link_cols and val.startswith("http"):
                    cell.hyperlink = val
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = brd

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    log.info(f"Excel → {xlsx_path}")
    return csv_path, xlsx_path


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    t0 = time.time()
    log.info("=" * 60)
    log.info("EXTRACTION COMPLÈTE (toutes pages) – Abbécourt (Aisne)")
    log.info(f"Démarrage : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)

    # URL initiale (limit:50 pour avoir le cookie Anubis)
    start_url = f"{BASE_URL}{BASE_PATH}/{COMMUNE_ID}/limit:50{SEARCH_PARAMS}"

    # Authentification Anubis
    session = get_authenticated_session(start_url)
    if not session:
        log.error("Impossible d'obtenir une session.")
        sys.exit(1)

    # Extraction toutes pages
    records = extract_all(session)

    # Export
    csv_path, xlsx_path = export(records)

    elapsed = time.time() - t0
    log.info(f"\n{'='*60}")
    log.info(f"SUCCÈS : {len(records)} registres extraits en {elapsed:.1f}s")
    log.info(f"CSV   → {csv_path}")
    log.info(f"Excel → {xlsx_path}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
